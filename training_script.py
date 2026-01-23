import os
import pandas as pd
import numpy as np
import joblib
import torch

from sentence_transformers import SentenceTransformer
from sklearn.ensemble import RandomForestClassifier
from sklearn.multioutput import MultiOutputClassifier
from sklearn.pipeline import Pipeline
from sklearn.base import BaseEstimator, TransformerMixin

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# ==============================================================================
# PATH & PERFORMANCE CONFIG
# ==============================================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DATASET_PATH = os.path.join(BASE_DIR, "datasets", "training_data.xlsx")
MODEL_DIR = os.path.join(BASE_DIR, "models")
os.makedirs(MODEL_DIR, exist_ok=True)

MODEL_NAME = "BAAI/bge-m3"

torch.set_num_threads(8)
os.environ["OMP_NUM_THREADS"] = "8"
os.environ["MKL_NUM_THREADS"] = "8"

AI_BATCH_SIZE = 256

# ==============================================================================
# PRELOAD SENTENCE TRANSFORMER (ONLINE, STABLE)
# ==============================================================================
print(f"🔄 Loading SentenceTransformer model: {MODEL_NAME}")
EMBEDDER = SentenceTransformer(MODEL_NAME, device="cpu")
print("✅ SentenceTransformer loaded")

# ==============================================================================
# DATA CONSTANTS (UNCHANGED)
# ==============================================================================
SHEET_DAILY = "PROCESS (DEFECT)"
SHEET_MONTHLY = "PROCESS (OZ,MS,IH)"

HEADER_KEYWORD = "PROC_DETAIL_E"

AI_CONTEXT_COLS = [
    "SYMPTOM_DESCRIPTION_E",
    "PROC_DETAIL_E",
    "ASC_REMARK_E",
]

TARGETS_DEFECT = ["Defect1", "Defect2", "Defect3"]
TARGETS_CATEGORY = ["SVC TYPE", "DETAIL REASON"]

FINAL_COLUMNS_ORDER = [
    "DATA_TYPE",
    "RCPT_NO_ORD_NO",
    "CLOSE_DT_RTN_DT",
    "SALES_MODEL_SUFFIX",
    "SERIAL_NO",
    "PARTS_DESC1",
    "PARTS_DESC2",
    "PARTS_DESC3",
    "PROC_DETAIL_E",
    "ASC_REMARK_E",
]

# ==============================================================================
# HEADER DETECTION (USED!)
# ==============================================================================
def find_header_index(file_path, sheet_name, keyword):
    df_temp = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=40)

    if isinstance(df_temp, dict):
        df_temp = list(df_temp.values())[0]

    for idx, row in df_temp.iterrows():
        row_text = row.astype(str).str.upper().tolist()
        if any(keyword.upper() in cell for cell in row_text):
            return idx

    return 0


# ==============================================================================
# AI INPUT PREPARATION (USED!)
# ==============================================================================
def prepare_ai_input(df):
    for col in AI_CONTEXT_COLS:
        if col not in df.columns:
            df[col] = ""

    df["AI_INPUT_COMBINED"] = (
        df["SYMPTOM_DESCRIPTION_E"].fillna("") + " "
        + df["PROC_DETAIL_E"].fillna("") + " "
        + df["ASC_REMARK_E"].fillna("")
    ).str.strip()

    return df


# ==============================================================================
# EXCEL STYLING (USED!)
# ==============================================================================
def apply_excel_styling(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_align

    ws.freeze_panes = "A2"
    wb.save(file_path)


# ==============================================================================
# EMBEDDER WRAPPER (PIPELINE SAFE)
# ==============================================================================
class BertEmbedder(BaseEstimator, TransformerMixin):
    def fit(self, X, y=None):
        return self

    def transform(self, X):
        return EMBEDDER.encode(
            list(X),
            batch_size=AI_BATCH_SIZE,
            show_progress_bar=False,
        )


# ==============================================================================
# TRAINING FUNCTION (USED BY API)
# ==============================================================================
def train_ai_advanced(enable_cleansing=False, progress_callback=None):
    def report(pct, msg):
        print(f"[{pct}%] {msg}")
        if progress_callback:
            progress_callback(pct, msg)

    # ------------------------------------------------------------------
    # LOAD DEFECT DATA
    # ------------------------------------------------------------------
    report(5, "Detecting header (DEFECT sheet)...")
    defect_header = find_header_index(DATASET_PATH, SHEET_DAILY, HEADER_KEYWORD)

    df_defect = pd.read_excel(
        DATASET_PATH,
        sheet_name=SHEET_DAILY,
        header=defect_header,
    )

    df_defect.columns = df_defect.columns.astype(str).str.strip()
    df_defect = prepare_ai_input(df_defect)

    missing = [c for c in TARGETS_DEFECT if c not in df_defect.columns]
    if missing:
        raise ValueError(f"Missing defect columns: {missing}")

    # ------------------------------------------------------------------
    # TRAIN DEFECT MODEL
    # ------------------------------------------------------------------
    report(25, "Training DEFECT model...")
    pipe_defect = Pipeline([
        ("embedder", BertEmbedder()),
        ("clf", MultiOutputClassifier(
            RandomForestClassifier(
                n_estimators=300,
                max_depth=25,
                n_jobs=8,
                random_state=42
            )
        ))
    ])

    pipe_defect.fit(
        df_defect["AI_INPUT_COMBINED"],
        df_defect[TARGETS_DEFECT].fillna("-"),
    )

    joblib.dump(pipe_defect, os.path.join(MODEL_DIR, "model_defect.pkl"))

    # ------------------------------------------------------------------
    # LOAD CATEGORY DATA
    # ------------------------------------------------------------------
    report(60, "Detecting header (CATEGORY sheet)...")
    cat_header = find_header_index(DATASET_PATH, SHEET_MONTHLY, HEADER_KEYWORD)

    df_cat = pd.read_excel(
        DATASET_PATH,
        sheet_name=SHEET_MONTHLY,
        header=cat_header,
    )

    df_cat.columns = df_cat.columns.astype(str).str.strip()
    df_cat = prepare_ai_input(df_cat)

    missing_cat = [c for c in TARGETS_CATEGORY if c not in df_cat.columns]
    if missing_cat:
        raise ValueError(f"Missing category columns: {missing_cat}")

    # ------------------------------------------------------------------
    # TRAIN CATEGORY MODEL
    # ------------------------------------------------------------------
    report(80, "Training CATEGORY model...")
    pipe_cat = Pipeline([
        ("embedder", BertEmbedder()),
        ("clf", MultiOutputClassifier(
            RandomForestClassifier(
                n_estimators=300,
                max_depth=25,
                n_jobs=8,
                random_state=42
            )
        ))
    ])

    pipe_cat.fit(
        df_cat["AI_INPUT_COMBINED"],
        df_cat[TARGETS_CATEGORY].fillna("-"),
    )

    joblib.dump(pipe_cat, os.path.join(MODEL_DIR, "model_oz.pkl"))

    report(100, "Training completed successfully 🎉")


# ==============================================================================
# PREDICTION + EXPORT (USED!)
# ==============================================================================
def predict_excel_process(input_path, output_path):
    pipe_defect = joblib.load(os.path.join(MODEL_DIR, "model_defect.pkl"))
    pipe_cat = joblib.load(os.path.join(MODEL_DIR, "model_oz.pkl"))

    df = pd.read_excel(input_path)
    df = prepare_ai_input(df)

    defect_pred = pipe_defect.predict(df["AI_INPUT_COMBINED"])
    cat_pred = pipe_cat.predict(df["AI_INPUT_COMBINED"])

    for i, col in enumerate(TARGETS_DEFECT):
        df[col] = defect_pred[:, i]

    for i, col in enumerate(TARGETS_CATEGORY):
        df[col] = cat_pred[:, i]

    df = df[[c for c in FINAL_COLUMNS_ORDER if c in df.columns] + TARGETS_DEFECT + TARGETS_CATEGORY]
    df.to_excel(output_path, index=False)

    apply_excel_styling(output_path)

    return output_path
