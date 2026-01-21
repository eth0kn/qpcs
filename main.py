from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import joblib
import io
import os
import shutil
import re
import threading
import time

# --- EXCEL STYLING LIBRARIES ---
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# --- AI & NLP LIBRARIES ---
from sentence_transformers import SentenceTransformer
from sklearn.base import BaseEstimator, TransformerMixin

# ==============================================================================
# 1. AI CLASS DEFINITION
# ==============================================================================
class BertEmbedder(BaseEstimator, TransformerMixin):
    def __init__(self, model_name):
        self.model_name = model_name
        self.model = None
    def fit(self, X, y=None):
        if self.model is None: self.model = SentenceTransformer(self.model_name)
        return self
    def transform(self, X):
        if self.model is None: self.model = SentenceTransformer(self.model_name)
        if hasattr(X, 'tolist'): sentences = X.tolist()
        else: sentences = X
        # Naikkan batch size karena RAM 16GB cukup
        return self.model.encode(sentences, batch_size=64, show_progress_bar=False)

# ==============================================================================
# 2. SERVER CONFIGURATION
# ==============================================================================
app = FastAPI(title="QPCS AI System API", version="13.0 (Fixed Output & Performance)")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

MODEL_DIR = 'models/'
DATASET_DIR = 'datasets/'
MODEL_DEFECT_PATH = os.path.join(MODEL_DIR, 'model_defect.pkl')
MODEL_OZ_PATH = os.path.join(MODEL_DIR, 'model_oz.pkl')

ai_models = {}
TRAINING_STATE = {"is_running": False, "progress": 0, "message": "Idle"}

# ==============================================================================
# 3. HELPER FUNCTIONS
# ==============================================================================
def load_system_resources():
    print("🚀 [SYSTEM START] Loading AI Models into RAM...")
    try:
        if os.path.exists(MODEL_DEFECT_PATH) and os.path.exists(MODEL_OZ_PATH):
            ai_models['defect'] = joblib.load(MODEL_DEFECT_PATH)
            ai_models['oz'] = joblib.load(MODEL_OZ_PATH)
            print("   ✅ AI Models: LOADED")
        else:
            print("   ⚠️ AI Models: NOT FOUND.")
    except Exception as e:
        print(f"   ❌ AI Models Error: {e}")

load_system_resources()

def clean_text_deep(text):
    if not isinstance(text, str): return ""
    text = text.lower().strip()
    text = re.sub(r'\b(pend_reason|tech_remark|asc_remark|pending|remark)\b', ' ', text)
    text = re.sub(r'\b(set ok|job done|replaced)\b', ' ', text)
    text = re.sub(r'[\:\-\_\|\=\[\]]', ' ', text)
    text = re.sub(r'\b(?=\w*\d)(?=\w*[a-z])\w{7,}\b', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def is_valid_text(text):
    # Filter diperlonggar agar tidak membuang data valid
    s = str(text).strip()
    if not s or s.lower() in ["nan", "null"]: return False
    return True

def filter_output_columns(df, input_col_name):
    # Kolom wajib yang harus ada di output
    required_cols = ['DATA_TYPE', 'RCPT_NO_ORD_NO', 'CLOSE_DT_RTN_DT', 'SALES_MODEL_SUFFIX', 'SERIAL_NO', 'PARTS_DESC1', 'PARTS_DESC2', 'PARTS_DESC3', 'PROC_DETAIL_E', 'ASC_REMARK_E']
    
    # Buat DataFrame baru
    df_clean = pd.DataFrame()
    
    # Copy kolom yang ada, isi kosong jika tidak ada
    for col in required_cols:
        if col in df.columns:
            df_clean[col] = df[col]
        else:
            df_clean[col] = "" # Isi string kosong biar rapi
            
    # Pastikan input column ada
    if 'PROC_DETAIL_E' not in df_clean.columns and input_col_name in df.columns:
        df_clean['PROC_DETAIL_E'] = df[input_col_name]
        
    return df_clean

# --- STYLING FUNCTION ---
def apply_excel_styling(ws, report_type, header_row_num, start_col_idx):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Conditional Colors
    fill_oz = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_ih = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_ms = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    svc_col_index = None

    # Style Header
    for row in ws.iter_rows(min_row=header_row_num, max_row=header_row_num, min_col=start_col_idx):
        for cell in row:
            if cell.value:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                if report_type == 'monthly' and str(cell.value).strip() == 'SVC TYPE':
                    svc_col_index = cell.column

    # Style Data
    for row in ws.iter_rows(min_row=header_row_num + 1, min_col=start_col_idx):
        for cell in row:
            # Cek apakah cell ini punya header di atasnya (berarti bagian tabel)
            header_cell = ws.cell(row=header_row_num, column=cell.column)
            if header_cell.value: 
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                
                if report_type == 'monthly' and svc_col_index and cell.column == svc_col_index:
                    val = str(cell.value).strip().upper()
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    if val == 'OZ': 
                        cell.fill = fill_oz
                        cell.font = Font(color="006100")
                    elif val == 'IH': 
                        cell.fill = fill_ih
                        cell.font = Font(color="9C0006")
                    elif val == 'MS': 
                        cell.fill = fill_ms
                        cell.font = Font(color="9C6500")

    # Auto Width
    for column in ws.columns:
        if column[0].column < start_col_idx: continue
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 60: adjusted_width = 60
        ws.column_dimensions[column_letter].width = adjusted_width

# ==============================================================================
# 4. PREDICTION ENDPOINT
# ==============================================================================
@app.post("/predict")
async def predict_excel(
    file: UploadFile = File(...),
    report_type: str = Query(..., description="Select: 'daily' or 'monthly'"),
    enable_cleansing: bool = Query(True)
):
    if not file.filename.endswith(('.xlsx', '.xls')): raise HTTPException(400, "Invalid file format.")

    try:
        contents = await file.read()
        df_raw = pd.read_excel(io.BytesIO(contents), engine='openpyxl')
        
        # 1. Cari Kolom Text
        input_col = 'PROC_DETAIL_E'
        if input_col not in df_raw.columns:
             candidates = [c for c in df_raw.columns if 'detail' in str(c).lower()]
             if candidates: input_col = candidates[0]
        
        # 2. Filter Kolom
        df_final = filter_output_columns(df_raw, input_col)
        
        # 3. Text Processing
        X_temp = df_final['PROC_DETAIL_E'].fillna("").astype(str)
        if enable_cleansing: X_temp = X_temp.apply(clean_text_deep)
        
        # Pastikan kita memprediksi SEMUA baris agar output sejajar, 
        # baris kosong akan dapat hasil default "-"
        valid_mask = X_temp.apply(is_valid_text)
        
        # Ambil data valid untuk diprediksi
        X_pred = X_temp[valid_mask].tolist()

        # 4. Prediksi
        if report_type == 'daily':
            # Init kolom kosong
            df_final['Defect1'] = "-"
            df_final['Defect2'] = "-"
            df_final['Defect3'] = "-"
            
            if len(X_pred) > 0 and 'defect' in ai_models:
                 pred_def = ai_models['defect'].predict(X_pred)
                 df_final.loc[valid_mask, 'Defect1'] = pred_def[:, 0]
                 df_final.loc[valid_mask, 'Defect2'] = pred_def[:, 1]
                 df_final.loc[valid_mask, 'Defect3'] = pred_def[:, 2]
            
            sheet_name = 'Defect Classification'
            start_row = 2   # Excel Row 3
            start_col = 1   # Excel Col B
            header_title = "DEFECT CLASSIFICATION (DAILY 1X PER DAY)"
            header_cell = "B2"
            
        elif report_type == 'monthly':
            df_final['Defect1'] = "-"
            df_final['Defect2'] = "-"
            df_final['Defect3'] = "-"
            df_final['SVC TYPE'] = "-"
            df_final['DETAIL REASON'] = "-"
            
            if len(X_pred) > 0:
                if 'defect' in ai_models:
                    pred_def = ai_models['defect'].predict(X_pred)
                    df_final.loc[valid_mask, 'Defect1'] = pred_def[:, 0]
                    df_final.loc[valid_mask, 'Defect2'] = pred_def[:, 1]
                    df_final.loc[valid_mask, 'Defect3'] = pred_def[:, 2]
                
                if 'oz' in ai_models:
                    pred_oz = ai_models['oz'].predict(X_pred)
                    df_final.loc[valid_mask, 'SVC TYPE'] = pred_oz[:, 0]
                    df_final.loc[valid_mask, 'DETAIL REASON'] = pred_oz[:, 1]
            
            sheet_name = 'OZ,MS,IH CATEGORY'
            start_row = 1   # Excel Row 2
            start_col = 0   # Excel Col A
            header_title = "OZ/MS/IH CATEGORY (MONTHLY 1X PER MONTH)"
            header_cell = "A1"

        # 5. Export dengan Layout Fix
        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            # Tulis data
            df_final.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row, startcol=start_col)
            
            worksheet = writer.sheets[sheet_name]
            
            # Tulis Judul
            title_font = Font(bold=True, size=12)
            worksheet[header_cell] = header_title
            worksheet[header_cell].font = title_font
            
            # Terapkan Styling
            apply_excel_styling(worksheet, report_type, header_row_num=(start_row + 1), start_col_idx=(start_col + 1))
        
        output_buffer.seek(0)
        prefix = "DAILY_" if report_type == 'daily' else "MONTHLY_"
        filename = f"{prefix}RESULT_{file.filename}"
        
        return StreamingResponse(
            output_buffer, 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"Error during prediction: {str(e)}")
        # Jangan raise 500 jika masalah data, tapi log ke console server
        raise HTTPException(500, f"Server Process Error: {str(e)}")

# ==============================================================================
# 5. TRAINING LOGIC
# ==============================================================================
def run_training_process(clean_bool):
    global TRAINING_STATE
    TRAINING_STATE["is_running"] = True
    TRAINING_STATE["progress"] = 0
    TRAINING_STATE["message"] = "Initializing..."
    try:
        from training_script import train_ai_advanced
        def update_progress(pct, msg):
            TRAINING_STATE["progress"] = pct
            TRAINING_STATE["message"] = msg
        train_ai_advanced(enable_cleansing=clean_bool, progress_callback=update_progress)
        load_system_resources()
    except Exception as e:
        TRAINING_STATE["message"] = f"Error: {str(e)}"
        TRAINING_STATE["progress"] = 0
    finally:
        TRAINING_STATE["is_running"] = False

@app.post("/train")
async def start_training(file: UploadFile = File(...), enable_cleansing: bool = Query(True)):
    global TRAINING_STATE
    if TRAINING_STATE["is_running"]: raise HTTPException(400, "Training in progress.")
    os.makedirs(DATASET_DIR, exist_ok=True)
    file_location = os.path.join(DATASET_DIR, "training_data.xlsx")
    try:
        with open(file_location, "wb+") as fo: shutil.copyfileobj(file.file, fo)
    except PermissionError:
        raise HTTPException(400, "File is locked/open in Excel. Close it first.")
    thread = threading.Thread(target=run_training_process, args=(enable_cleansing,))
    thread.start()
    return {"status": "started"}

@app.get("/train/status")
async def get_training_status():
    return TRAINING_STATE

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)